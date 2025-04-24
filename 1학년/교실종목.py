import tkinter as tk
from tkinter import messagebox
import configparser
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_SETTINGS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "setting.ini")
RESULTS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "결과", "교실.xlsx")

def load_base_settings():
    config = configparser.ConfigParser()
    try:
        config.read(BASE_SETTINGS_PATH, encoding="utf-8")
        class_count = int(config.get("기본설정", "class"))
        ini_path = os.path.join(os.path.dirname(BASE_SETTINGS_PATH), "설정", "교실.ini")
        if not os.path.exists(ini_path):
            raise FileNotFoundError(f"설정 경로가 존재하지 않습니다: {ini_path}")
        return class_count, ini_path
    except Exception as e:
        messagebox.showerror("오류", f"설정 파일을 불러오는 중 문제가 발생했습니다.\n{str(e)}")
        exit()

def load_event_settings(path):
    config = configparser.ConfigParser()
    try:
        config.read(path, encoding="utf-8")
        events = []
        for section in config.sections():
            if section.startswith("종목"):
                name = config.get(section, "이름", fallback="Unknown")
                high_first = config.get(section, "높은값우선", fallback="True").strip().lower() in ["true", "yes", "1"]
                scores = {
                    int(k.replace("등", "")): int(v)
                    for k, v in config.items(section) if "등" in k
                }
                if name and scores:
                    events.append((name, high_first, scores))
        return events
    except Exception as e:
        messagebox.showerror("오류", f"교실 설정 파일을 불러오는 중 문제가 발생했습니다.\n{str(e)}")
        exit()

def load_existing_inputs(filepath):
    if not os.path.exists(filepath):
        return {}
    df = pd.read_excel(filepath)
    if "반" not in df.columns:
        return {}
    existing_data = {}
    for col in df.columns:
        if col.endswith("입력값"):
            event = col.replace(" 입력값", "")
            temp = df[["반", col]].copy()
            temp.set_index("반", inplace=True)
            existing_data[event] = temp[col].to_dict()
    return existing_data

def convert_to_rank(scores, high_first):
    indexed = list(enumerate(scores))
    indexed.sort(key=lambda x: x[1], reverse=high_first)
    result = [0] * len(scores)
    rank = 1
    for i, (idx, val) in enumerate(indexed):
        if i > 0 and val == indexed[i - 1][1]:
            result[idx] = result[indexed[i - 1][0]]
        else:
            result[idx] = rank
        rank += 1
    return result

def calculate_scores():
    try:
        result = {"반": [f"{i+1}반" for i in range(class_count)]}
        for event_name, high_first, score_map in events:
            raw = []
            for i in range(class_count):
                try:
                    val = float(entries[event_name][i].get())
                    raw.append(val)
                except:
                    messagebox.showerror("오류", f"{i+1}반의 {event_name} 점수가 올바르지 않습니다.")
                    return
            ranks = convert_to_rank(raw, high_first)
            final = [score_map.get(r, 0) for r in ranks]
            result[f"{event_name} 입력값"] = [str(v) for v in raw]
            result[f"{event_name} 등수"] = [f"{r}등" for r in ranks]
            result[f"{event_name} 점수"] = [f"{s}점" for s in final]
        df = pd.DataFrame(result)
        df.to_excel(RESULTS_PATH, index=False)

        wb = load_workbook(RESULTS_PATH)
        ws = wb.active
        for col_cells in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col_cells if cell.value)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = max_len + 12
        wb.save(RESULTS_PATH)

        messagebox.showinfo("완료", "점수 계산이 완료되었습니다!")
    except Exception as e:
        messagebox.showerror("오류", f"계산 중 오류 발생: {e}")

def create_gui():
    root = tk.Tk()
    root.title("교실 종목 점수 입력")
    global entries
    entries = {}
    existing_inputs = load_existing_inputs(RESULTS_PATH)

    tk.Label(root, text="반 번호", width=10).grid(row=0, column=0)
    for idx, (event_name, _, _) in enumerate(events):
        tk.Label(root, text=event_name, width=15).grid(row=0, column=idx+1)

    for i in range(class_count):
        ban = f"{i+1}반"
        tk.Label(root, text=ban, width=10).grid(row=i+1, column=0)
        for idx, (event_name, _, _) in enumerate(events):
            entry = tk.Entry(root, width=15)
            entry.grid(row=i+1, column=idx+1)
            if event_name not in entries:
                entries[event_name] = []
            entries[event_name].append(entry)
            if event_name in existing_inputs and ban in existing_inputs[event_name]:
                val = existing_inputs[event_name][ban]
                if pd.notna(val):
                    entry.insert(0, str(val))

    tk.Button(root, text="점수 계산", command=calculate_scores, width=20).grid(row=class_count+2, column=0, columnspan=len(events)+1, pady=10)
    root.mainloop()

if __name__ == "__main__":
    class_count, event_path = load_base_settings()
    events = load_event_settings(event_path)
    create_gui()
