import tkinter as tk
from tkinter import messagebox
import configparser
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_SETTINGS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "setting.ini")

def load_base_settings():
    config = configparser.ConfigParser()
    try:
        config.read(BASE_SETTINGS_PATH, encoding="utf-8")
        class_count = int(config.get("기본설정", "class"))
        path = config.get("기본설정", "PATH")
        path = path.rstrip("/\\")
        path = os.path.join(path, "설정", "운동장.ini")
        
        if not os.path.exists(path):
            raise FileNotFoundError(f"설정 경로가 존재하지 않습니다: {path}")
        
        return class_count, path
    except Exception as e:
        messagebox.showerror("오류", f"설정 파일을 불러오는 중 문제가 발생했습니다.\n{str(e)}")
        exit()

def load_event_settings(path):
    config = configparser.ConfigParser()
    try:
        config.read(path, encoding="utf-8")
        events = []

        for section in config.sections():
            if section.startswith("종목"):
                event_name = config.get(section, "이름", fallback="Unknown")
                high_value_priority_str = config.get(section, "높은값우선", fallback="True").strip().lower()
                
                high_value_priority = high_value_priority_str in ["true", "yes", "1"]
                
                scores = {}
                for key, value in config.items(section):
                    if "등" in key:
                        try:
                            rank = int(key.replace("등", ""))
                            score = int(value)
                            scores[rank] = score
                        except ValueError:
                            messagebox.showerror("오류", f"점수 값이 숫자가 아닙니다: {key} = {value}")
                            exit()
                
                if event_name != "Unknown" and scores:
                    events.append((event_name, high_value_priority, scores))
        
        if not events:
            raise ValueError("운동장 설정 파일에 유효한 종목이 없습니다.")
        
        return events
    except Exception as e:
        messagebox.showerror("오류", f"운동장 설정 파일을 불러오는 중 문제가 발생했습니다.\n{str(e)}")
        exit()

def convert_to_rank(scores, high_value_priority):
    indexed_scores = list(enumerate(scores))
    indexed_scores.sort(key=lambda x: x[1], reverse=high_value_priority)

    ranks = [0] * len(scores)
    current_rank = 1
    for i, (idx, score) in enumerate(indexed_scores):
        if i > 0 and score == indexed_scores[i - 1][1]:
            ranks[idx] = ranks[indexed_scores[i - 1][0]]
        else:
            ranks[idx] = current_rank
        current_rank += 1
    return ranks

def calculate_scores():
    try:
        results = {"반": [f"{i+1}반" for i in range(class_count)]}

        for event_name, high_value_priority, event_scores in events:
            raw_scores = []

            for i in range(class_count):
                try:
                    score = float(entries[event_name][i].get())
                    raw_scores.append(score)
                except ValueError:
                    messagebox.showerror("오류", f"{i+1}반의 {event_name} 점수를 올바르게 입력하세요!")
                    return
            
            ranks = convert_to_rank(raw_scores, high_value_priority)

            final_scores = [event_scores.get(rank, 0) for rank in ranks]

            input_col = f"{event_name} 입력값"
            score_col = f"{event_name} 점수"
            rank_col = f"{event_name} 등수"
            results[input_col] = [str(s) for s in raw_scores]
            results[rank_col] = [f"{r}등" for r in ranks]
            results[score_col] = [f"{s}점" for s in final_scores]

        save_results(results)
    except Exception as e:
        messagebox.showerror("오류", f"점수 계산 중 문제가 발생했습니다: {str(e)}")

def save_results(results):
    df = pd.DataFrame(results)
    df.to_excel(RESULTS_PATH, index=False)

    wb = load_workbook(RESULTS_PATH)
    ws = wb.active
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 8  # 기존보다 넓게 설정

    wb.save(RESULTS_PATH)
    messagebox.showinfo("완료", "점수 계산이 완료되었습니다!")

def create_gui():
    root = tk.Tk()
    root.title("운동장 종목 점수 입력")

    global entries
    entries = {}

    title_frame = tk.Frame(root)
    title_frame.pack(pady=10)
    tk.Label(title_frame, text="반 번호", width=10, borderwidth=2, relief="solid").grid(row=0, column=0)

    for idx, (event_name, _, _) in enumerate(events):
        tk.Label(title_frame, text=event_name, width=15, borderwidth=2, relief="solid").grid(row=0, column=idx+1)

    input_frame = tk.Frame(root)
    input_frame.pack()

    for i in range(class_count):
        tk.Label(input_frame, text=f"{i+1}반", width=10, borderwidth=2, relief="solid").grid(row=i, column=0)
        for idx, (event_name, _, _) in enumerate(events):
            if event_name not in entries:
                entries[event_name] = []
            entry = tk.Entry(input_frame, width=10, borderwidth=2, relief="solid")
            entry.grid(row=i, column=idx+1, padx=5, pady=5)
            entries[event_name].append(entry)

    calc_button = tk.Button(root, text="점수 계산", command=calculate_scores, width=20, height=2)
    calc_button.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    try:
        class_count, event_settings_path = load_base_settings()
        event_settings_path = os.path.abspath(event_settings_path)
        events = load_event_settings(event_settings_path)
        RESULTS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "결과", "운동장.xlsx")
        create_gui()
    except Exception as e:
        messagebox.showerror("오류", f"프로그램 실행 중 오류가 발생했습니다: {str(e)}")
