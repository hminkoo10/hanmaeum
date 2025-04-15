import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 결과 폴더 및 파일 설정
RESULT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), "결과")
FILES = ["공원강당.xlsx", "교실.xlsx", "운동장.xlsx"]
FINAL_FILE = "최종.xlsx"

# "100점" -> 100.0
def extract_numeric(value):
    try:
        number = re.findall(r"[\d.]+", str(value))
        return float(number[0]) if number else 0
    except:
        return 0

# 파일에서 점수 컬럼만 추출
def load_and_process(file_path):
    df = pd.read_excel(file_path)
    print(f"\n📄 [{os.path.basename(file_path)}] 열 이름 목록:", list(df.columns))

    if df.empty:
        raise ValueError(f"❌ 파일이 비어있습니다: {file_path}")

    if not any(col.strip() == "반" for col in df.columns):
        raise ValueError(f"❌ '반' 열이 없습니다: {file_path}")

    # 실제 반 컬럼 찾기
    actual_ban_col = [col for col in df.columns if col.strip() == "반"][0]

    # 점수 컬럼만 선택
    score_cols = [col for col in df.columns if col.endswith("점수")]
    out_df = df[[actual_ban_col] + score_cols].copy()
    out_df.rename(columns={actual_ban_col: "반"}, inplace=True)

    # 점수 문자열을 숫자로 변환
    for col in score_cols:
        out_df[col] = out_df[col].apply(extract_numeric)

    return out_df

# 메인 로직
def main():
    merged_df = None

    for file in FILES:
        file_path = os.path.join(RESULT_FOLDER, file)
        if not os.path.exists(file_path):
            print(f"⚠️ 파일 없음: {file_path}")
            continue

        partial = load_and_process(file_path)

        if merged_df is None:
            merged_df = partial
        else:
            merged_df = pd.merge(merged_df, partial, on="반", how="outer")

    if merged_df is None or merged_df.empty:
        raise ValueError("❌ 처리할 유효한 데이터가 없습니다. 모든 파일이 비어 있거나 누락되었습니다.")

    merged_df.fillna(0, inplace=True)

    # 총점 계산
    score_cols = [col for col in merged_df.columns if col.endswith("점수")]
    merged_df["총점"] = merged_df[score_cols].sum(axis=1)

    # 순위 계산 ("X위" 형식)
    merged_df["순위"] = merged_df["총점"].rank(method="min", ascending=False).astype(int).astype(str) + "위"

    # 반 정렬 (1반~10반)
    merged_df["__반번호"] = merged_df["반"].apply(lambda x: int(re.findall(r"\d+", str(x))[0]) if re.findall(r"\d+", str(x)) else 999)
    merged_df.sort_values("__반번호", inplace=True)
    merged_df.drop(columns="__반번호", inplace=True)

    # 모든 값 문자열로 변환
    for col in merged_df.columns:
        merged_df[col] = merged_df[col].apply(lambda x: f"{x}" if not pd.isnull(x) else "")

    # 저장
    final_path = os.path.join(RESULT_FOLDER, FINAL_FILE)
    merged_df.to_excel(final_path, index=False)

    # 셀 너비 자동 조정 (길게 표시되도록)
    wb = load_workbook(final_path)
    ws = wb.active
    for col_cells in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col_cells if cell.value)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max_len + 12  # 여유롭게

    wb.save(final_path)
    print(f"\n✅ 최종 결과 저장 완료: {final_path}")

if __name__ == "__main__":
    main()
    input("\n🔚 작업이 완료되었습니다. Enter 키를 누르면 종료됩니다...")